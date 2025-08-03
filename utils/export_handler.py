"""
Group Export Handler - Handles exporting group participants to JSON, Excel, and CSV formats
"""

import json
import logging
import os
import csv
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class GroupExportHandler:
    """Handles exporting group participants to various formats"""
    
    def __init__(self, export_dir: str = "static/exports"):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    def export_group_participants(
        self, 
        participants: List[Dict[str, Any]], 
        group_name: str,
        session_name: str
    ) -> Dict[str, Any]:
        """
        Export group participants to JSON and Excel formats
        
        Args:
            participants: List of participant details
            group_name: Name of the WhatsApp group
            session_name: WhatsApp session name
            
        Returns:
            Dictionary with export results and file paths
        """
        try:
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_group_name = self._sanitize_filename(group_name)
            base_filename = f"{safe_group_name}_participants_{timestamp}"
            
            # Prepare data for export
            export_data = self._prepare_export_data(participants, group_name)
            
            # Export to JSON
            json_path = self._export_to_json(export_data, base_filename)
            
            # Export to Excel
            excel_path = self._export_to_excel(export_data, base_filename, group_name)
            
            # Export to CSV
            csv_path = self._export_to_csv(export_data, base_filename)
            
            # Generate download URLs (relative to static directory)
            json_url = f"/static/exports/{os.path.basename(json_path)}"
            excel_url = f"/static/exports/{os.path.basename(excel_path)}"
            csv_url = f"/static/exports/{os.path.basename(csv_path)}"
            
            return {
                "success": True,
                "json_url": json_url.replace('/static', ''),  # Remove /static prefix
                "excel_url": excel_url.replace('/static', ''),  # Remove /static prefix
                "csv_url": csv_url.replace('/static', ''),  # Remove /static prefix
                "participant_count": len(participants),
                "group_name": group_name,
                "exported_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Failed to export group participants: {str(e)}")
            raise
    
    def _prepare_export_data(self, participants: List[Dict], group_name: str) -> Dict[str, Any]:
        """Prepare participant data for export"""
        
        # Participants are already in the correct format from get_group_participants_details
        # Just ensure all fields are present with defaults
        processed_participants = []
        
        for participant in participants:
            # Ensure all fields have a value (use existing or default)
            processed = {
                'phone_number': participant.get('phone_number', ''),
                'formatted_phone': participant.get('formatted_phone', ''),
                'country_code': participant.get('country_code', ''),
                'country_name': participant.get('country_name', 'Unknown'),
                'saved_name': participant.get('saved_name', ''),
                'public_name': participant.get('public_name', ''),
                'is_my_contact': participant.get('is_my_contact', False),
                'is_business': participant.get('is_business', False),
                'is_blocked': participant.get('is_blocked', False),
                'is_admin': participant.get('is_admin', False),
                'is_super_admin': participant.get('is_super_admin', False),
                'labels': participant.get('labels', ''),
                'last_msg_text': participant.get('last_msg_text', ''),
                'last_msg_date': participant.get('last_msg_date', ''),
                'last_msg_type': participant.get('last_msg_type', ''),
                'last_msg_status': participant.get('last_msg_status', '')
            }
            
            processed_participants.append(processed)
        
        return {
            "group_name": group_name,
            "export_date": datetime.now().isoformat(),
            "total_participants": len(participants),
            "participants": processed_participants
        }
    
    def _export_to_json(self, data: Dict[str, Any], filename: str) -> str:
        """Export data to JSON format"""
        json_path = os.path.join(self.export_dir, f"{filename}.json")
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Exported to JSON: {json_path}")
        return json_path
    
    def _export_to_excel(self, data: Dict[str, Any], filename: str, group_name: str) -> str:
        """Export data to Excel format with formatting"""
        excel_path = os.path.join(self.export_dir, f"{filename}.xlsx")
        
        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Participants"
        
        # Add column headers directly at row 1
        headers = [
            'phone_number', 'formatted_phone', 'country_code', 'country_name',
            'saved_name', 'public_name', 'is_my_contact', 'is_business', 
            'is_blocked', 'is_admin', 'is_super_admin', 'labels',
            'last_msg_text', 'last_msg_date', 'last_msg_type', 'last_msg_status'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write participant data starting from row 2
        for row_idx, participant in enumerate(data['participants'], 2):
            ws.cell(row=row_idx, column=1, value=participant['phone_number'])
            ws.cell(row=row_idx, column=2, value=participant['formatted_phone'])
            ws.cell(row=row_idx, column=3, value=participant['country_code'])
            ws.cell(row=row_idx, column=4, value=participant['country_name'])
            ws.cell(row=row_idx, column=5, value=participant['saved_name'])
            ws.cell(row=row_idx, column=6, value=participant['public_name'])
            ws.cell(row=row_idx, column=7, value='Yes' if participant['is_my_contact'] else 'No')
            ws.cell(row=row_idx, column=8, value='Yes' if participant['is_business'] else 'No')
            ws.cell(row=row_idx, column=9, value='Yes' if participant['is_blocked'] else 'No')
            ws.cell(row=row_idx, column=10, value='Yes' if participant['is_admin'] else 'No')
            ws.cell(row=row_idx, column=11, value='Yes' if participant['is_super_admin'] else 'No')
            ws.cell(row=row_idx, column=12, value=participant['labels'])
            ws.cell(row=row_idx, column=13, value=participant['last_msg_text'])
            ws.cell(row=row_idx, column=14, value=participant['last_msg_date'])
            ws.cell(row=row_idx, column=15, value=participant['last_msg_type'])
            ws.cell(row=row_idx, column=16, value=participant['last_msg_status'])
            
            # Alternate row coloring
            if row_idx % 2 == 0:
                for col in range(1, 17):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(excel_path)
        logger.info(f"Exported to Excel: {excel_path}")
        return excel_path
    
    def _export_to_csv(self, data: Dict[str, Any], filename: str) -> str:
        """Export data to CSV format"""
        csv_path = os.path.join(self.export_dir, f"{filename}.csv")
        
        # Define headers using snake_case to match processor expectations
        headers = [
            'phone_number', 'formatted_phone', 'country_code', 'country_name',
            'saved_name', 'public_name', 'is_my_contact', 'is_business', 
            'is_blocked', 'is_admin', 'is_super_admin', 'labels',
            'last_msg_text', 'last_msg_date', 'last_msg_type', 'last_msg_status'
        ]
        
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            
            # Write participant data
            for participant in data['participants']:
                # Create row with exact field names
                row = {
                    'phone_number': participant.get('formatted_phone', ''),  # Use formatted_phone which has the + prefix
                    'formatted_phone': participant.get('formatted_phone', ''),
                    'country_code': f"+{participant.get('country_code', '')}" if participant.get('country_code', '') else '',
                    'country_name': participant.get('country_name', ''),
                    'saved_name': participant.get('saved_name', ''),
                    'public_name': participant.get('public_name', ''),
                    'is_my_contact': 'true' if participant.get('is_my_contact') else 'false',
                    'is_business': 'true' if participant.get('is_business') else 'false',
                    'is_blocked': 'true' if participant.get('is_blocked') else 'false',
                    'is_admin': 'true' if participant.get('is_admin') else 'false',
                    'is_super_admin': 'true' if participant.get('is_super_admin') else 'false',
                    'labels': participant.get('labels', ''),
                    'last_msg_text': participant.get('last_msg_text', ''),
                    'last_msg_date': participant.get('last_msg_date', ''),
                    'last_msg_type': participant.get('last_msg_type', ''),
                    'last_msg_status': participant.get('last_msg_status', '')
                }
                writer.writerow(row)
        
        logger.info(f"Exported to CSV: {csv_path}")
        return csv_path
    
    def _sanitize_filename(self, filename: str) -> str:
        """Sanitize filename to remove invalid characters"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename[:50]  # Limit length
    
    def _format_timestamp(self, timestamp: Any) -> str:
        """Format timestamp to readable date string"""
        if not timestamp:
            return ''
        
        try:
            # Handle both Unix timestamp and datetime strings
            if isinstance(timestamp, (int, float)):
                dt = datetime.fromtimestamp(timestamp)
            else:
                dt = datetime.fromisoformat(str(timestamp))
            
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return str(timestamp)
